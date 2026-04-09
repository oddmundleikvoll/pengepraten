#!/usr/bin/env python3
"""
Pengepraten.no - Budsjettmal Generator
Lager en komplett Excel-budsjettmal med 4 faner
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from datetime import datetime

# Opprett workbook
wb = Workbook()

# Farger
GRONN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GRONN_MORK = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
ROD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ROD_MORK = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
GRA = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BLA = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GUL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HVIT = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Fonter
font_tittel = Font(bold=True, size=16, color="1F4E79")
font_undertittel = Font(bold=True, size=12, color="1F4E79")
font_bold = Font(bold=True)
font_hvit = Font(bold=True, color="FFFFFF")
font_liten = Font(size=9, italic=True)

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Hjelpefunksjon for å justere kolonnebredde
def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================================
# FANE 1: MÅNEDLIG BUDSJETT
# ============================================================================
ws1 = wb.active
ws1.title = "Månedlig budsjett"

# Header
ws1['A1'] = "PENGEPRATEN.NO - MÅNEDLIG BUDSJETT"
ws1['A1'].font = font_tittel
ws1.merge_cells('A1:C1')

ws1['A2'] = "Måned:"
ws1['B2'] = datetime.now().strftime("%B %Y")
ws1['B2'].font = font_bold
ws1['A2'].font = font_bold

ws1['A4'] = "INNTEKTER"
ws1['A4'].font = font_undertittel
ws1['A4'].fill = BLA
ws1['A4'].border = thin_border
ws1['B4'].fill = BLA
ws1['B4'].border = thin_border
ws1['C4'].fill = BLA
ws1['C4'].border = thin_border

inntekter = [
    ("Lønn (netto)", "Din månedslønn etter skatt"),
    ("Bonus/ekstrainntekt", "Ekstra inntekter denne måneden"),
    ("Barnetrygd", ""),
    ("Stønad/bidrag", "Barnebidrag, bostøtte, etc."),
    ("Annen inntekt", "Leieinntekt, utbytte, etc."),
]

row = 5
for navn, kommentar in inntekter:
    ws1[f'A{row}'] = navn
    ws1[f'B{row}'] = 0
    ws1[f'B{row}'].number_format = '#,##0 kr'
    if kommentar:
        ws1[f'A{row}'].comment = Comment(kommentar, "Pengepraten")
    row += 1

ws1[f'A{row}'] = "SUM INNTEKTER"
ws1[f'A{row}'].font = font_bold
ws1[f'B{row}'] = f"=SUM(B5:B{row-1})"
ws1[f'B{row}'].font = font_bold
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].fill = GRONN
inntekt_sum_row = row

# Faste utgifter
row += 2
ws1[f'A{row}'] = "FASTE UTGIFTER"
ws1[f'A{row}'].font = font_undertittel
ws1[f'A{row}'].fill = ROD
ws1[f'B{row}'].fill = ROD
ws1[f'C{row}'].fill = ROD
ws1[f'A{row}'].border = thin_border
ws1[f'B{row}'].border = thin_border
ws1[f'C{row}'].border = thin_border

faste_utgifter = [
    ("Husleie/huslån", "Månedlig leie eller terminbeløp"),
    ("Strøm", "Gjennomsnittlig strømregning"),
    ("Internett/TV", "Bredbånd og TV-pakke"),
    ("Forsikringer", "Hus, bil, liv, reise"),
    ("Mobilabonnement", "Alle mobilabonnement i husholdningen"),
    ("Strømmetjenester", "Netflix, Spotify, etc."),
    ("Aviser/medlemskap", "Avisabonnement, treningssenter, etc."),
    ("Barnehage/SFO", "Månedlig utgift"),
    ("Annen fast utgift", ""),
]

row += 1
start_faste = row
for navn, kommentar in faste_utgifter:
    ws1[f'A{row}'] = navn
    ws1[f'B{row}'] = 0
    ws1[f'B{row}'].number_format = '#,##0 kr'
    if kommentar:
        ws1[f'A{row}'].comment = Comment(kommentar, "Pengepraten")
    row += 1

ws1[f'A{row}'] = "SUM FASTE UTGIFTER"
ws1[f'A{row}'].font = font_bold
ws1[f'B{row}'] = f"=SUM(B{start_faste}:B{row-1})"
ws1[f'B{row}'].font = font_bold
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].fill = ROD
faste_sum_row = row

# Variable utgifter
row += 2
ws1[f'A{row}'] = "VARIABLE UTGIFTER"
ws1[f'A{row}'].font = font_undertittel
ws1[f'A{row}'].fill = ROD
ws1[f'B{row}'].fill = ROD
ws1[f'C{row}'].fill = ROD
ws1[f'A{row}'].border = thin_border
ws1[f'B{row}'].border = thin_border
ws1[f'C{row}'].border = thin_border

variable_utgifter = [
    ("Mat og drikke", "Dagligvarer, takeaway, restaurant"),
    ("Transport", "Bensin, billettar, vedlikehold"),
    ("Helse", "Legemidler, egenandeler, briller"),
    ("Klær og sko", ""),
    ("Personlig pleie", "Frisør, kosmetikk"),
    ("Fritid/underholdning", "Kino, hobby, sport"),
    ("Gaver", "Bursdager, jul, etc."),
    ("Annet", "Uforutsette utgifter"),
]

row += 1
start_variable = row
for navn, kommentar in variable_utgifter:
    ws1[f'A{row}'] = navn
    ws1[f'B{row}'] = 0
    ws1[f'B{row}'].number_format = '#,##0 kr'
    if kommentar:
        ws1[f'A{row}'].comment = Comment(kommentar, "Pengepraten")
    row += 1

ws1[f'A{row}'] = "SUM VARIABLE UTGIFTER"
ws1[f'A{row}'].font = font_bold
ws1[f'B{row}'] = f"=SUM(B{start_variable}:B{row-1})"
ws1[f'B{row}'].font = font_bold
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].fill = ROD
variable_sum_row = row

# Sparing
row += 2
ws1[f'A{row}'] = "SPARING"
ws1[f'A{row}'].font = font_undertittel
ws1[f'A{row}'].fill = GUL
ws1[f'B{row}'].fill = GUL
ws1[f'C{row}'].fill = GUL
ws1[f'A{row}'].border = thin_border
ws1[f'B{row}'].border = thin_border
ws1[f'C{row}'].border = thin_border

sparing = [
    ("Bufferkonto", "Nødfond - mål: 3-6 måneders utgifter"),
    ("Langsiktig sparing", "Aksjer, fond, pensjon"),
    ("Spare til ferie", ""),
    ("Spare til bil/annet", ""),
]

row += 1
start_sparing = row
for navn, kommentar in sparing:
    ws1[f'A{row}'] = navn
    ws1[f'B{row}'] = 0
    ws1[f'B{row}'].number_format = '#,##0 kr'
    if kommentar:
        ws1[f'A{row}'].comment = Comment(kommentar, "Pengepraten")
    row += 1

ws1[f'A{row}'] = "SUM SPARING"
ws1[f'A{row}'].font = font_bold
ws1[f'B{row}'] = f"=SUM(B{start_sparing}:B{row-1})"
ws1[f'B{row}'].font = font_bold
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].fill = GUL
sparing_sum_row = row

# Oppsummering
row += 3
ws1[f'A{row}'] = "OPPSUMMERING"
ws1[f'A{row}'].font = font_tittel

row += 1
ws1[f'A{row}'] = "Totale utgifter:"
ws1[f'B{row}'] = f"=B{faste_sum_row}+B{variable_sum_row}+B{sparing_sum_row}"
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].font = font_bold

row += 1
ws1[f'A{row}'] = "Disponibelt (Inntekt - Utgifter):"
ws1[f'A{row}'].font = font_bold
ws1[f'B{row}'] = f"=B{inntekt_sum_row}-B{row-1}"
ws1[f'B{row}'].font = Font(bold=True, size=14)
ws1[f'B{row}'].number_format = '#,##0 kr'
ws1[f'B{row}'].comment = Comment("Grønn = positivt (bra!), Rød = negativt (sjekk utgiftene)", "Pengepraten")

# Conditional formatting for disponibelt
from openpyxl.formatting.rule import CellIsRule
ws1.conditional_formatting.add(f'B{row}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=GRONN))
ws1.conditional_formatting.add(f'B{row}',
    CellIsRule(operator='lessThan', formula=['0'], fill=ROD))
ws1.conditional_formatting.add(f'B{row}',
    CellIsRule(operator='equal', formula=['0'], fill=GUL))

# Kolonnebredder
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 40

# ============================================================================
# FANE 2: ÅRSBUDSJETT
# ============================================================================
ws2 = wb.create_sheet("Årsbudsjett")

ws2['A1'] = "PENGEPRATEN.NO - ÅRSBUDSJETT"
ws2['A1'].font = font_tittel
ws2.merge_cells('A1:N1')

ws2['A2'] = "År:"
ws2['B2'] = datetime.now().year
ws2['B2'].font = font_bold

# Header rad
headers = ["Kategori", "Jan", "Feb", "Mar", "Apr", "Mai", "Jun", 
           "Jul", "Aug", "Sep", "Okt", "Nov", "Des", "Totalt"]
row = 4
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col, value=header)
    cell.font = font_bold
    cell.fill = BLA
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Kategorier
kategorier = [
    "Lønn og inntekter",
    "Faste utgifter",
    "Variable utgifter", 
    "Sparing",
    "Resultat (Inntekt - Utgifter)"
]

row = 5
for kat in kategorier:
    ws2.cell(row=row, column=1, value=kat).font = font_bold
    if "Resultat" in kat:
        ws2.cell(row=row, column=1).fill = GUL
    row += 1

# Formler for månedlige summer (bruker dummy-data, bruker fyller inn)
# Resultat-rad
resultat_row = 9
for month in range(2, 14):  # Jan-Des kolonner
    col_letter = get_column_letter(month)
    ws2.cell(row=resultat_row, column=month, value=f"={col_letter}5-{col_letter}6-{col_letter}7-{col_letter}8")
    ws2.cell(row=resultat_row, column=month).number_format = '#,##0 kr'
    ws2.cell(row=resultat_row, column=month).font = font_bold

# Totalt-kolonne
for r in range(5, 10):
    ws2.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
    ws2.cell(row=r, column=14).number_format = '#,##0 kr'
    ws2.cell(row=r, column=14).font = font_bold
    ws2.cell(row=r, column=14).fill = GRA

# Årlig mål-seksjon
row = 12
ws2[f'A{row}'] = "ÅRLIGE MÅL"
ws2[f'A{row}'].font = font_undertittel
ws2[f'A{row}'].fill = GRONN_MORK
ws2[f'A{row}'].font = font_hvit
ws2.merge_cells(f'A{row}:C{row}')

row += 1
ws2[f'A{row}'] = "Sparemål for året:"
ws2[f'B{row}'] = 50000
ws2[f'B{row}'].number_format = '#,##0 kr'
ws2[f'C{row}'] = "(Endre til ditt mål)"
ws2[f'C{row}'].font = font_liten

row += 1
ws2[f'A{row}'] = "Faktisk sparing (hittil):"
ws2[f'B{row}'] = "=SUM(B8:M8)"
ws2[f'B{row}'].number_format = '#,##0 kr'

row += 1
ws2[f'A{row}'] = "Fremdrift:"
ws2[f'B{row}'] = f"=B{row-1}/B{row-2}"
ws2[f'B{row}'].number_format = '0%'
ws2[f'B{row}'].font = font_bold

# Månedlig gjennomsnitt
row += 2
ws2[f'A{row}'] = "SNITT PER MÅNED"
ws2[f'A{row}'].font = font_undertittel
ws2[f'A{row}'].fill = BLA
ws2.merge_cells(f'A{row}:C{row}')

snitt_kategorier = ["Inntekter", "Faste utgifter", "Variable utgifter", "Sparing", "Disponibelt"]
snitt_rader = [5, 6, 7, 8, 9]

row += 1
for kat, r in zip(snitt_kategorier, snitt_rader):
    ws2[f'A{row}'] = f"{kat}:"
    ws2[f'B{row}'] = f"=N{r}/12"
    ws2[f'B{row}'].number_format = '#,##0 kr'
    row += 1

# Kolonnebredder
ws2.column_dimensions['A'].width = 25
for col in range(2, 15):
    ws2.column_dimensions[get_column_letter(col)].width = 12

# ============================================================================
# FANE 3: GJELDSBREMS
# ============================================================================
ws3 = wb.create_sheet("Gjeldsbrems")

ws3['A1'] = "PENGEPRATEN.NO - GJELDSBREMS"
ws3['A1'].font = font_tittel
ws3.merge_cells('A1:G1')

ws3['A2'] = "Få oversikt over gjelden din og se hvor fort du kan bli gjeldfri!"
ws3['A2'].font = font_liten
ws3.merge_cells('A2:G2')

# Header
headers = ["Gjeldstype", "Resterende beløp", "Årlig rente (%)", "Min. betaling/mnd", 
           "Ekstra innbetaling", "Tid til nedbetaling", "Totalt betalt"]
row = 4
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col, value=header)
    cell.font = font_bold
    cell.fill = ROD_MORK
    cell.font = font_hvit
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Gjeldseksempler
gjeldstyper = [
    ("Kredittkort 1", 25000, 25, 750),
    ("Kredittkort 2", 15000, 22, 500),
    ("Forbrukslån", 100000, 15, 2500),
    ("Billån", 150000, 7, 3500),
    ("Studielån", 200000, 5, 1500),
    ("Annet", 0, 0, 0),
]

row = 5
for navn, belop, rente, minbet in gjeldstyper:
    ws3.cell(row=row, column=1, value=navn)
    ws3.cell(row=row, column=2, value=belop).number_format = '#,##0 kr'
    ws3.cell(row=row, column=3, value=rente/100 if rente > 1 else rente).number_format = '0.00%'
    ws3.cell(row=row, column=4, value=minbet).number_format = '#,##0 kr'
    ws3.cell(row=row, column=5, value=0).number_format = '#,##0 kr'
    
    # Formel for tid til nedbetaling (forenklet)
    col_b = get_column_letter(2)
    col_d = get_column_letter(4)
    col_e = get_column_letter(5)
    ws3.cell(row=row, column=6, value=f"=IF(OR({col_b}{row}=0,({col_d}{row}+{col_e}{row})=0),\"Betalt!\",ROUND({col_b}{row}/({col_d}{row}+{col_e}{row}),1))")
    
    # Totalt betalt (forenklet estimat)
    ws3.cell(row=row, column=7, value=f"=IF({col_b}{row}=0,0,{col_b}{row}*1.2)").number_format = '#,##0 kr'
    
    for col in range(1, 8):
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

# Oppsummering
row += 1
ws3[f'A{row}'] = "TOTALT"
ws3[f'A{row}'].font = font_bold
ws3[f'A{row}'].fill = ROD

ws3[f'B{row}'] = f"=SUM(B5:B{row-2})"
ws3[f'B{row}'].font = font_bold
ws3[f'B{row}'].number_format = '#,##0 kr'
ws3[f'B{row}'].fill = ROD

ws3[f'D{row}'] = f"=SUM(D5:D{row-2})"
ws3[f'D{row}'].font = font_bold
ws3[f'D{row}'].number_format = '#,##0 kr'
ws3[f'D{row}'].fill = ROD

ws3[f'E{row}'] = f"=SUM(E5:E{row-2})"
ws3[f'E{row}'].font = font_bold
ws3[f'E{row}'].number_format = '#,##0 kr'
ws3[f'E{row}'].fill = ROD

ws3[f'G{row}'] = f"=SUM(G5:G{row-2})"
ws3[f'G{row}'].font = font_bold
ws3[f'G{row}'].number_format = '#,##0 kr'
ws3[f'G{row}'].fill = ROD

# Strategi-seksjon
row += 3
ws3[f'A{row}'] = "NEDBETALINGSTRATEGI"
ws3[f'A{row}'].font = font_undertittel
ws3[f'A{row}'].fill = GRONN_MORK
ws3[f'A{row}'].font = font_hvit
ws3.merge_cells(f'A{row}:G{row}')

row += 1
ws3[f'A{row}'] = "Avalanchesmetoden (anbefalt):"
ws3[f'A{row}'].font = font_bold
ws3[f'B{row}'] = "Betal minimum på alt, ekstra på høyeste rente først"
ws3.merge_cells(f'B{row}:G{row}')

row += 1
ws3[f'A{row}'] = "Snøballmetoden:"
ws3[f'A{row}'].font = font_bold
ws3[f'B{row}'] = "Betal minimum på alt, ekstra på minste lån først (motivasjon!)"
ws3.merge_cells(f'B{row}:G{row}')

row += 2
ws3[f'A{row}'] = "Hvor mye raskere blir du gjeldfri med ekstra innbetalinger?"
ws3[f'A{row}'].font = font_bold
ws3[f'A{row}'].fill = GUL
ws3.merge_cells(f'A{row}:G{row}')

row += 1
ws3[f'A{row}'] = "Månedlig ekstra innbetaling:"
ws3[f'B{row}'] = 1000
ws3[f'B{row}'].number_format = '#,##0 kr'
ws3[f'C{row}'] = "Tid spart:"
ws3[f'D{row}'] = "~6-12 måneder"
ws3[f'D{row}'].font = font_bold
ws3[f'D{row}'].fill = GRONN

# Kolonnebredder
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 18

# ============================================================================
# FANE 4: SPAREPLAN
# ============================================================================
ws4 = wb.create_sheet("Spareplan")

ws4['A1'] = "PENGEPRATEN.NO - SPAREPLAN"
ws4['A1'].font = font_tittel
ws4.merge_cells('A1:F1')

ws4['A2'] = "Sett deg mål, følg fremdrift, nå drømmene dine!"
ws4['A2'].font = font_liten
ws4.merge_cells('A2:F2')

# Header
headers = ["Sparemål", "Målbeløp", "Spart hittil", "Månedlig sparing", "Fremdrift (%)", "Est. ferdig"]
row = 4
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col, value=header)
    cell.font = font_bold
    cell.fill = GRONN_MORK
    cell.font = font_hvit
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Sparekategorier
sparemal = [
    ("Bufferkonto (nødfond)", 50000, 0, 2000),
    ("Feriekonto", 30000, 5000, 1500),
    ("Ny bil", 150000, 20000, 3000),
    ("Egenkapital bolig", 300000, 50000, 5000),
    ("Langsiktig sparing", 100000, 10000, 1000),
    ("Annet mål", 0, 0, 0),
]

row = 5
for navn, mal, spart, mnd in sparemal:
    ws4.cell(row=row, column=1, value=navn)
    ws4.cell(row=row, column=2, value=mal).number_format = '#,##0 kr'
    ws4.cell(row=row, column=3, value=spart).number_format = '#,##0 kr'
    ws4.cell(row=row, column=4, value=mnd).number_format = '#,##0 kr'
    
    # Fremdrift %
    ws4.cell(row=row, column=5, value=f"=IF(B{row}=0,0,C{row}/B{row})")
    ws4.cell(row=row, column=5).number_format = '0%'
    
    # Estimert ferdig
    ws4.cell(row=row, column=6, value=f"=IF(OR(D{row}=0,B{row}=0),\"-\",DATE(YEAR(TODAY()),MONTH(TODAY())+ROUNDUP((B{row}-C{row})/D{row},0),1))")
    ws4.cell(row=row, column=6).number_format = 'mmm YYYY'
    
    for col in range(1, 7):
        ws4.cell(row=row, column=col).border = thin_border
    row += 1

# Oppsummering
row += 1
ws4[f'A{row}'] = "TOTALT"
ws4[f'A{row}'].font = font_bold
ws4[f'A{row}'].fill = GRONN

ws4[f'B{row}'] = f"=SUM(B5:B{row-2})"
ws4[f'B{row}'].font = font_bold
ws4[f'B{row}'].number_format = '#,##0 kr'
ws4[f'B{row}'].fill = GRONN

ws4[f'C{row}'] = f"=SUM(C5:C{row-2})"
ws4[f'C{row}'].font = font_bold
ws4[f'C{row}'].number_format = '#,##0 kr'
ws4[f'C{row}'].fill = GRONN

ws4[f'D{row}'] = f"=SUM(D5:D{row-2})"
ws4[f'D{row}'].font = font_bold
ws4[f'D{row}'].number_format = '#,##0 kr'
ws4[f'D{row}'].fill = GRONN

ws4[f'E{row}'] = f"=IF(B{row}=0,0,C{row}/B{row})"
ws4[f'E{row}'].font = font_bold
ws4[f'E{row}'].number_format = '0%'
ws4[f'E{row}'].fill = GRONN

# Tips-seksjon
row += 3
ws4[f'A{row}'] = "TIPS FOR Å NÅ MÅLENE"
ws4[f'A{row}'].font = font_undertittel
ws4[f'A{row}'].fill = BLA
ws4.merge_cells(f'A{row}:F{row}')

tips = [
    "🎯 Automatiser sparingen - sett opp fast trekk på lønningsdagen",
    "📈 Prioriter bufferkonto først (3-6 måneders utgifter)",
    "🏦 Bruk høyrentekonto til kortsiktige mål",
    "📊 Gjennomgå spareplanen hver måned",
    "🎉 Feire når du når delmål!"
]

row += 1
for tip in tips:
    ws4[f'A{row}'] = tip
    ws4.merge_cells(f'A{row}:F{row}')
    row += 1

# Kolonnebredder
ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 15

# ============================================================================
# Lagre filen
# ============================================================================
output_path = "/home/o/.openclaw/workspace/projects/norsk-penger/emails/pengepraten-budsjettmal.xlsx"
wb.save(output_path)
print(f"✅ Budsjettmal lagret til: {output_path}")

# Lag også CSV-versjon (kun første ark)
import csv

csv_path = "/home/o/.openclaw/workspace/projects/norsk-penger/emails/pengepraten-budsjettmal.csv"
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    for row in ws1.iter_rows(values_only=True):
        writer.writerow(row)
print(f"✅ CSV-versjon lagret til: {csv_path}")
